import itertools
import os
import random
import re
import time
import warnings
from tkinter import filedialog

from openpyxl import Workbook, load_workbook

warnings.filterwarnings("ignore")

### export list of column in excel file
### input (excel file, column head)
### output (column data)
def Func_read_excel_column(excel_filename, column_name):
    excel = load_workbook(filename=excel_filename)
    excel_sheetnames = excel.get_sheet_names()
    excel_data = excel.get_sheet_by_name(excel_sheetnames[0])
    txt_file = []
    for i in range(excel_data.max_row):
        row_current = ""
        for j in range(excel_data.max_column):
            row_current += str(excel_data.cell(row=(i + 1), column=(j + 1)).value)
            row_current += "\t"
        txt_file.append(row_current)
    txt_data = [line.split("\t") for line in txt_file]
    for i in range(len(txt_data[0])):
        if txt_data[0][i] == column_name:
            column_target = i
    return [txt_data[i][column_target] for i in range(1, len(txt_data))]


### export list of column in txt file
### input (txt file, column header)
### output (column data)
def Func_read_txt_column(txt_filename, column_name):
    txt_file = [line.strip() for line in open(txt_filename).read().splitlines()]
    txt_data = [line.split("\t") for line in txt_file]
    for i in range(len(txt_data[0])):
        if txt_data[0][i] == column_name:
            column_target = i
    return [txt_data[i][column_target] for i in range(1, len(txt_data))]


## program
## program

Input_result = filedialog.askopenfilename(
    initialdir=(os.getcwd()), filetypes=[("input data", ".txt")], title=("input data")
)
print(Input_result)
Input_list_pro = "D-Va-step4-protein-filter$.xlsx"

List_pro = Func_read_excel_column(Input_list_pro, "Proteins")
List_pro_ID = Func_read_txt_column(Input_result, "Proteins")

List_pro_ID.insert(0, "Proteins")

with open(Input_result.split(".")[0] + "_pro_filter.txt", "w") as output:
    line_count = 0
    for line in open(Input_result):
        output_check = False
        if line_count == 0:
            output_check = True
        for i in range(len(List_pro)):
            if List_pro[i] in List_pro_ID[line_count]:
                output_check = True
        if output_check:
            output.write(line)
        line_count += 1
